"""
Utilidades para exportar plazos judiciales a diferentes formatos.
Incluye exportación a PDF e iCalendar.
"""

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from datetime import datetime
import weasyprint
from ics import Calendar, Event
from .plazos import formatear_fecha_chilena, es_plazo_urgente


def exportar_pdf(plazos, titulo="Calendario de Plazos Judiciales"):
    """
    Exporta una lista de plazos judiciales a PDF.
    
    Args:
        plazos: QuerySet de plazos judiciales
        titulo: Título del documento PDF
    
    Returns:
        HttpResponse con el PDF generado
    """
    # Preparar datos para el template
    context = {
        'plazos': plazos,
        'titulo': titulo,
        'fecha_generacion': timezone.now(),
        'total_plazos': plazos.count(),
    }
    
    # Renderizar HTML
    html_string = render_to_string('plazos/export/pdf_template.html', context)
    
    # Generar PDF
    pdf_file = weasyprint.HTML(string=html_string).write_pdf()
    
    # Crear respuesta HTTP
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="plazos_judiciales_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response


def exportar_ics(plazos, titulo="Plazos Judiciales"):
    """
    Exporta una lista de plazos judiciales a formato iCalendar (.ics).
    
    Args:
        plazos: QuerySet de plazos judiciales
        titulo: Título del calendario
    
    Returns:
        HttpResponse con el archivo .ics generado
    """
    # Crear calendario
    cal = Calendar()
    
    for plazo in plazos:
        if not plazo.fecha_vencimiento:
            continue
        
        # Crear evento
        evento = Event()
        evento.uid = f"plazo-{plazo.id}@calendario-judicial.local"
        evento.name = f"{plazo.get_tipo_documento_display()} - {plazo.rut_causa}"
        evento.begin = plazo.fecha_vencimiento
        evento.end = plazo.fecha_vencimiento
        
        # Descripción del evento
        descripcion = f"""
Tipo de Documento: {plazo.get_tipo_documento_display()}
Procedimiento: {plazo.get_procedimiento_display()}
Días de Plazo: {plazo.dias_plazo} ({plazo.get_tipo_dia_display()})
Fecha de Inicio: {formatear_fecha_chilena(plazo.fecha_inicio)}
Estado: {plazo.get_estado_display()}
Clave del Cliente: {plazo.clave_cliente}
"""
        
        if plazo.observaciones:
            descripcion += f"\nObservaciones: {plazo.observaciones}"
        
        evento.description = descripcion.strip()
        
        # Categorías
        evento.categories = [plazo.get_procedimiento_display(), plazo.get_tipo_documento_display()]
        
        # Prioridad según urgencia
        if es_plazo_urgente(plazo.fecha_vencimiento):
            evento.priority = 1  # Alta prioridad
        elif plazo.estado == 'vencido':
            evento.priority = 0  # Muy alta prioridad
        else:
            evento.priority = 5  # Normal
        
        # Agregar evento al calendario
        cal.events.add(evento)
    
    # Generar contenido iCalendar
    ics_content = str(cal)
    
    # Crear respuesta HTTP
    response = HttpResponse(ics_content, content_type='text/calendar; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="plazos_judiciales_{datetime.now().strftime("%Y%m%d_%H%M%S")}.ics"'
    
    return response


def exportar_excel(plazos, titulo="Plazos Judiciales"):
    """
    Exporta una lista de plazos judiciales a Excel.
    Nota: Requiere openpyxl o xlsxwriter
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl es requerido para exportar a Excel. Instálelo con: pip install openpyxl")
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plazos Judiciales"
    
    # Encabezados
    headers = [
        'ID', 'Tipo Documento', 'Procedimiento', 'Días Plazo', 'Tipo Día',
        'Fecha Inicio', 'Fecha Vencimiento', 'RUT Causa', 'Clave Cliente',
        'Estado', 'Días Restantes', 'Observaciones'
    ]
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Escribir datos
    for row, plazo in enumerate(plazos, 2):
        ws.cell(row=row, column=1, value=plazo.id)
        ws.cell(row=row, column=2, value=plazo.get_tipo_documento_display())
        ws.cell(row=row, column=3, value=plazo.get_procedimiento_display())
        ws.cell(row=row, column=4, value=plazo.dias_plazo)
        ws.cell(row=row, column=5, value=plazo.get_tipo_dia_display())
        ws.cell(row=row, column=6, value=plazo.fecha_inicio.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=7, value=plazo.fecha_vencimiento.strftime('%d/%m/%Y') if plazo.fecha_vencimiento else '')
        ws.cell(row=row, column=8, value=plazo.rut_causa)
        ws.cell(row=row, column=9, value=plazo.clave_cliente)
        ws.cell(row=row, column=10, value=plazo.get_estado_display())
        ws.cell(row=row, column=11, value=plazo.dias_restantes)
        ws.cell(row=row, column=12, value=plazo.observaciones or '')
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="plazos_judiciales_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Guardar workbook en la respuesta
    wb.save(response)
    
    return response
